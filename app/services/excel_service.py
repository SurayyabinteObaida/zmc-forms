"""
excel_service.py  (V2)
Generates Excel exports for Flexo and Gravure form types.
  - Flexo: Serial Wise Data + Raw Data sheets (unchanged from V1).
  - Gravure: Header sheet + Roll Rows sheet.
  - generate_batch_excel() dispatches based on form_type_code.
"""
import io
import json
import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from app.models import FormType

# ── Config persistence (Flexo only — Gravure columns are fixed) ───────────────
CONFIG_PATH = os.path.join(os.path.dirname(__file__), "export_config.json")

ALL_COLUMNS = [
    ("S. NO",             "serial",          6),
    ("Job Code",          "job_code",        14),
    ("PRINT DATE",        "date",            12),
    ("Job Name",          "job_name",        22),
    ("BAG SIZE",          "bag_size",        14),
    ("PRINTED FILM",      "material",        18),
    ("TUBE / SHEET",      "tube_sheet",      12),
    ("SIZE (INCHES)",     "web_size",        12),
    ("MIC",               "mic",             10),
    ("FILM SUPPLIER",     "supplier",        15),
    ("ORDER QTY",         "order_qty",       10),
    ("PLAIN WT",          "net_wt",          10),
    ("PRINTED WT",        "printed_reel_wt", 10),
    ("Printed Wastage",   "printed_waste",   12),
    ("PLAIN WASTAGE",     "plain_waste",     12),
    ("OPERATOR",          "operator",        12),
    ("SETTING TIME",      "setting_time",    12),
    ("START TIME",        "start_time",      10),
    ("END TIME",          "end_time",        10),
]

MANDATORY_KEYS = {"serial"}


def _default_config():
    return {col[1]: True for col in ALL_COLUMNS}


def load_export_config():
    if os.path.exists(CONFIG_PATH):
        try:
            with open(CONFIG_PATH) as f:
                saved = json.load(f)
            cfg = _default_config()
            cfg.update(saved)
            return cfg
        except Exception:
            pass
    return _default_config()


def save_export_config(config: dict):
    for k in MANDATORY_KEYS:
        config[k] = True
    with open(CONFIG_PATH, "w") as f:
        json.dump(config, f, indent=2)


def get_active_columns(config=None):
    if config is None:
        config = load_export_config()
    return [col for col in ALL_COLUMNS if config.get(col[1], True)]


# ── Shared Styles ─────────────────────────────────────────────────────────────
HEADER_FILL     = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT     = Font(color="FFFFFF", bold=True, size=10, name="Arial")
GRAVURE_HEADER_FILL = PatternFill(start_color="1a3a5c", end_color="1a3a5c", fill_type="solid")
ALT_ROW_FILL    = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
GRAVURE_ALT_FILL = PatternFill(start_color="E8F0FA", end_color="E8F0FA", fill_type="solid")
CORRECTION_FILL = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
THIN   = Side(style="thin")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _col(idx):
    return get_column_letter(idx)


def _apply_header(ws, headers, row=1, fill=None):
    fill = fill or HEADER_FILL
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=ci)
        c.value = h
        c.fill = fill
        c.font = HEADER_FONT
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDER
    ws.row_dimensions[row].height = 32


def _style_data_row(ws, row, ncols, alt, corrected_cols=None, alt_fill=None):
    alt_fill = alt_fill or ALT_ROW_FILL
    for ci in range(1, ncols + 1):
        c = ws.cell(row=row, column=ci)
        c.border = BORDER
        c.alignment = Alignment(vertical="center")
        if corrected_cols and ci in corrected_cols:
            c.fill = CORRECTION_FILL
        elif alt:
            c.fill = alt_fill


def _auto_col_width(ws):
    for col in ws.columns:
        w = max((len(str(c.value or "")) for c in col), default=10)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(w + 4, 40)


def _get_data(record):
    raw = getattr(record, "verified_data", None) \
       or getattr(record, "raw_extraction", None) \
       or "{}"
    return json.loads(raw)


# ── Flexo Export ──────────────────────────────────────────────────────────────

def _write_serial_wise(ws, records, corrections_map=None, config=None):
    corrections_map = corrections_map or {}
    active_cols = get_active_columns(config)
    headers = [c[0] for c in active_cols]
    ncols = len(headers)
    _apply_header(ws, headers, row=1)

    col_of = {c[1]: i + 1 for i, c in enumerate(active_cols)}

    for row_idx, record in enumerate(records, 1):
        excel_row = row_idx + 1
        data = _get_data(record)
        corrections = corrections_map.get(record.id, {})
        corrected_ci = {col_of[fk] for fk in corrections if fk in col_of}

        if "serial" in col_of:
            ws.cell(excel_row, col_of["serial"]).value = row_idx

        for label, dk, width in active_cols:
            if dk == "serial":
                continue
            ws.cell(excel_row, col_of[dk]).value = data.get(dk) or ""

        _style_data_row(ws, excel_row, ncols, row_idx % 2 == 0, corrected_ci)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{_col(ncols)}1"
    for i, (_, _, width) in enumerate(active_cols, 1):
        ws.column_dimensions[_col(i)].width = width


def _generate_flexo_excel(records, corrections_map=None):
    config = load_export_config()
    wb = Workbook()

    ws_serial = wb.active
    ws_serial.title = "Serial Wise Data"
    _write_serial_wise(ws_serial, records, corrections_map, config)

    ws_raw = wb.create_sheet("Raw Data")
    raw_headers = ["#", "File", "Saved At", "Date", "Job Name", "Job Code", "Operator",
                   "Material", "Supplier", "Web Size", "Mic", "Ink GSM", "Tube/Sheet",
                   "Bag Size", "Setting Time", "Start Time", "End Time",
                   "Plain Net Wt.", "Plain Waste", "Printed Net Wt.", "Printed Waste"]
    _apply_header(ws_raw, raw_headers)
    for ri, record in enumerate(records, 1):
        data = _get_data(record)
        saved_at = record.saved_at.strftime("%Y-%m-%d %H:%M") if record.saved_at else ""
        row = [
            ri, record.filename or "", saved_at,
            data.get("date", ""), data.get("job_name", ""), data.get("job_code", ""),
            data.get("operator", ""), data.get("material", ""), data.get("supplier", ""),
            data.get("web_size", ""), data.get("mic", ""), data.get("ink_gsm", ""),
            data.get("tube_sheet", ""), data.get("bag_size", ""),
            data.get("setting_time", ""), data.get("start_time", ""), data.get("end_time", ""),
            data.get("net_wt", ""), data.get("plain_waste", ""),
            data.get("printed_reel_wt", ""), data.get("printed_waste", ""),
        ]
        ws_raw.append(row)
        _style_data_row(ws_raw, ri + 1, len(raw_headers), ri % 2 == 0)
    _auto_col_width(ws_raw)
    ws_raw.freeze_panes = "A2"

    return wb


# ── Gravure Export ────────────────────────────────────────────────────────────

# Gravure header-level columns
GRAVURE_HEADER_COLS = [
    ("#",                  "serial"),
    ("File",               "filename"),
    ("Print Date",         "print_date"),
    ("Job Name",           "job_name"),
    ("Job Code",           "job_code"),
    ("Material",           "material"),
    ("Supplier",           "material_supplier"),
    ("Web Size & MIC",     "web_size_mic"),
    ("Plain Order Qty",    "plain_order_qty"),
    ("Cylinder Qty & #",   "cylinder_qty_number"),
    ("Cylinder L x Cir",   "cylinder_length_cir"),
    ("Speed",              "speed"),
    ("Operator",           "operator"),
    ("Color Man",          "color_man"),
    ("Ink GSM",            "ink_gsm"),
    ("Setting Time",       "setting_time"),
    ("Start Time",         "start_time"),
    ("End Time",           "end_time"),
    ("Plain Gross Wt.",    "plain_gross_wt"),
    ("Printed Gross Wt.",  "printed_gross_wt"),
    ("Plain Core Wt.",     "plain_core_wt"),
    ("Printed Core Wt.",   "printed_core_wt"),
    ("Plain Balance",      "plain_balance"),
    ("Plain Net Wt.",      "plain_net_wt"),
    ("Printed Net Wt.",    "printed_net_wt"),
    ("Total Mtr.",         "total_mtr"),
    ("Plain Waste",        "plain_waste"),
    ("Roll Waste",         "roll_waste"),
    ("Printed Waste",      "printed_waste"),
    ("Setting Waste",      "setting_waste"),
    ("Total Waste",        "total_waste"),
    ("Prepared By",        "prepared_by"),
    ("Supervisor",         "supervisor"),
    ("Remarks",            "remarks"),
]

# Gravure roll row columns
GRAVURE_ROLL_COLS = [
    ("Form #",              "form_serial"),
    ("Job Name",            "job_name"),
    ("Print Date",          "print_date"),
    ("RM #",                "rm_number"),
    ("Plain Roll Wt.",      "plain_roll_wt"),
    ("Plain Bal. Rejected", "plain_balance_rejected"),
    ("Plain Core Wt.",      "plain_core_wt"),
    ("Printed Roll #",      "printed_roll_number"),
    ("Printed Roll Wt.",    "printed_roll_wt"),
    ("Printed Core Wt.",    "printed_core_wt"),
    ("Meter",               "meter"),
    ("Remarks",             "remarks"),
]


def _generate_gravure_excel(records, corrections_map=None, header_cols=None):
    wb = Workbook()

    # Header columns come from the form's configured fields when available,
    # so the export matches the verify screen exactly. "#" and File are kept
    # as row identifiers. Falls back to the fixed list if no fields configured.
    if header_cols:
        cols = [("#", "serial"), ("File", "filename")] + list(header_cols)
    else:
        cols = GRAVURE_HEADER_COLS

    # ── Sheet 1: Header summary (one row per form) ────────────────────────────
    ws_hdr = wb.active
    ws_hdr.title = "Form Summary"
    hdr_labels = [c[0] for c in cols]
    _apply_header(ws_hdr, hdr_labels, fill=GRAVURE_HEADER_FILL)

    col_of_hdr = {c[1]: i + 1 for i, c in enumerate(cols)}

    for ri, record in enumerate(records, 1):
        excel_row = ri + 1
        data = _get_data(record)
        # Remove roll_rows from data dict so it doesn't pollute header fields
        data.pop("roll_rows", None)

        if "serial" in col_of_hdr:
            ws_hdr.cell(excel_row, col_of_hdr["serial"]).value = ri
        if "filename" in col_of_hdr:
            ws_hdr.cell(excel_row, col_of_hdr["filename"]).value = record.filename or ""

        for label, dk in cols:
            if dk in ("serial", "filename"):
                continue
            ws_hdr.cell(excel_row, col_of_hdr[dk]).value = data.get(dk, "") or ""

        _style_data_row(ws_hdr, excel_row, len(cols),
                        ri % 2 == 0, alt_fill=GRAVURE_ALT_FILL)

    ws_hdr.freeze_panes = "A2"
    ws_hdr.auto_filter.ref = f"A1:{_col(len(cols))}1"
    _auto_col_width(ws_hdr)

    # ── Sheet 2: Roll rows (one row per roll, linked back to form) ────────────
    ws_rolls = wb.create_sheet("Roll Detail")
    roll_labels = [c[0] for c in GRAVURE_ROLL_COLS]
    _apply_header(ws_rolls, roll_labels, fill=GRAVURE_HEADER_FILL)

    col_of_roll = {c[1]: i + 1 for i, c in enumerate(GRAVURE_ROLL_COLS)}
    roll_excel_row = 2

    for form_idx, record in enumerate(records, 1):
        data = _get_data(record)
        roll_rows = data.get("roll_rows", []) or []
        job_name = data.get("job_name", "")
        print_date = data.get("print_date", "")

        for row in roll_rows:
            if not any(v for v in row.values() if v is not None and v != ""):
                continue

            ws_rolls.cell(roll_excel_row, col_of_roll["form_serial"]).value = form_idx
            ws_rolls.cell(roll_excel_row, col_of_roll["job_name"]).value = job_name
            ws_rolls.cell(roll_excel_row, col_of_roll["print_date"]).value = print_date

            for label, dk in GRAVURE_ROLL_COLS:
                if dk in ("form_serial", "job_name", "print_date"):
                    continue
                ws_rolls.cell(roll_excel_row, col_of_roll[dk]).value = row.get(dk, "") or ""

            _style_data_row(ws_rolls, roll_excel_row, len(GRAVURE_ROLL_COLS),
                            roll_excel_row % 2 == 0, alt_fill=GRAVURE_ALT_FILL)
            roll_excel_row += 1

    ws_rolls.freeze_panes = "A2"
    ws_rolls.auto_filter.ref = f"A1:{_col(len(GRAVURE_ROLL_COLS))}1"
    _auto_col_width(ws_rolls)

    return wb


def _form_type_for_records(records):
    """Look up the FormType from the first record (all records in a batch share it)."""
    if not records:
        return None
    ft_id = getattr(records[0], "form_type_id", None)
    return FormType.query.get(ft_id) if ft_id else None


def _field_columns(form_type):
    """Header columns built from the form's enabled fields: [(label, key), ...].
    This is exactly what the verify screen shows, so the export matches it."""
    if not form_type:
        return None
    return [(f.label, f.key) for f in form_type.enabled_fields()]


# ── Public API ────────────────────────────────────────────────────────────────

def generate_batch_excel(batch_id, records, corrections_map=None, form_type_code=None):
    """
    Dispatch to the correct generator based on form_type_code.
    Falls back to flexo for backward compatibility.
    """
    form_type = _form_type_for_records(records)
    code = form_type_code or (form_type.code if form_type else None)
    if code == "F-PRD/01.1":
        wb = _generate_gravure_excel(records, corrections_map,
                                     header_cols=_field_columns(form_type))
    else:
        wb = _generate_flexo_excel(records, corrections_map)

    wb.properties.title = f"Batch {batch_id} — {datetime.now().strftime('%Y-%m-%d')}"
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def generate_excel(form_type, records, corrections_map=None):
    """Legacy single-form export. Dispatches by form_type.code."""
    code = form_type.code if hasattr(form_type, "code") else None
    return generate_batch_excel("export", records, corrections_map, form_type_code=code)


def get_column_config_for_ui():
    """Return Flexo columns with current enabled state for the settings UI."""
    config = load_export_config()
    return [
        {
            "key": col[1],
            "label": col[0],
            "enabled": config.get(col[1], True),
            "mandatory": col[1] in MANDATORY_KEYS,
        }
        for col in ALL_COLUMNS
    ]
#