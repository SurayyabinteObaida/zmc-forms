import json
from datetime import datetime, timezone
from flask import Blueprint, render_template, request, jsonify, send_file
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill
from app import db
from app.models import (FormType, ExtractedRecord, Submission,
                         FlexoPrintingRecord, GravurePrintingRecord)
from openpyxl.styles import Alignment, Border, Side
from openpyxl.utils import get_column_letter

verify_bp = Blueprint("verify", __name__)

FORM_MODEL_MAP = {
    "F-PRD/01.2": FlexoPrintingRecord,
    "F-PRD/01.1": GravurePrintingRecord,
}


@verify_bp.route("/<int:submission_id>")
def index(submission_id):
    submission = Submission.query.get_or_404(submission_id)
    records = ExtractedRecord.query.filter_by(submission_id=submission_id).all()

    enriched = []
    for rec in records:
        form_type = FormType.query.get(rec.form_type_id) if rec.form_type_id else None
        raw_data = json.loads(rec.raw_extraction or "{}")
        enabled_fields = form_type.enabled_fields() if form_type else []
        enriched.append({
            "record": rec,
            "form_type": form_type,
            "raw_data": raw_data,
            "enabled_fields": enabled_fields,
        })

    form_types_all = FormType.query.all()
    return render_template("verify/index.html", submission=submission, enriched=enriched, form_types=form_types_all)


@verify_bp.route("/record/<int:record_id>", methods=["GET"])
def get_record(record_id):
    rec = ExtractedRecord.query.get_or_404(record_id)
    form_type = FormType.query.get(rec.form_type_id) if rec.form_type_id else None
    data = json.loads(rec.raw_extraction or "{}")
    fields = [f.to_dict() for f in form_type.enabled_fields()] if form_type else []

    return jsonify({
        "record_id": rec.id,
        "filename": rec.filename,
        "form_code": form_type.code if form_type else None,
        "form_name": form_type.name if form_type else "Unknown",
        "confidence": rec.confidence,
        "status": rec.status,
        "error": rec.error_message,
        "fields": fields,
        "data": data,
        "image_path": rec.image_path,
    })


@verify_bp.route("/save/<int:record_id>", methods=["POST"])
def save_record(record_id):
    rec = ExtractedRecord.query.get_or_404(record_id)
    form_type = FormType.query.get(rec.form_type_id) if rec.form_type_id else None

    if not form_type:
        return jsonify({"error": "Unknown form type — cannot save"}), 400

    payload = request.get_json()
    verified_data = payload.get("data", {})

    rec.verified_data = json.dumps(verified_data)
    rec.status = "saved"
    rec.saved_at = datetime.now(timezone.utc)

    ModelClass = FORM_MODEL_MAP.get(form_type.code)
    if ModelClass:
        existing = ModelClass.query.filter_by(extracted_record_id=rec.id).first()
        if existing:
            db.session.delete(existing)
            db.session.flush()

        typed = ModelClass(extracted_record_id=rec.id)
        for key, value in verified_data.items():
            if hasattr(typed, key):
                setattr(typed, key, value if value != "" else None)
        db.session.add(typed)

    db.session.commit()
    return jsonify({"success": True, "record_id": rec.id, "status": rec.status})


@verify_bp.route("/save-batch/<int:submission_id>", methods=["POST"])
def save_batch(submission_id):
    submission = Submission.query.get_or_404(submission_id)
    payload = request.get_json()
    records_data = payload.get("records", {})

    saved_ids = []
    errors = []

    for rid_str, verified_data in records_data.items():
        rid = int(rid_str)
        rec = ExtractedRecord.query.get(rid)
        if not rec:
            errors.append(f"Record {rid} not found")
            continue

        form_type = FormType.query.get(rec.form_type_id) if rec.form_type_id else None
        if not form_type:
            errors.append(f"Record {rid}: unknown form type")
            continue

        rec.verified_data = json.dumps(verified_data)
        rec.status = "saved"
        rec.saved_at = datetime.now(timezone.utc)

        ModelClass = FORM_MODEL_MAP.get(form_type.code)
        if ModelClass:
            existing = ModelClass.query.filter_by(extracted_record_id=rec.id).first()
            if existing:
                db.session.delete(existing)
                db.session.flush()
            typed = ModelClass(extracted_record_id=rec.id)
            for key, value in verified_data.items():
                if hasattr(typed, key):
                    setattr(typed, key, value if value != "" else None)
            db.session.add(typed)

        saved_ids.append(rid)

    submission.status = "saved"
    db.session.commit()

    return jsonify({"success": True, "saved": saved_ids, "errors": errors})


@verify_bp.route("/update-form-type/<int:record_id>", methods=["POST"])
def update_form_type(record_id):
    rec = ExtractedRecord.query.get_or_404(record_id)
    payload = request.get_json()
    form_type_id = payload.get("form_type_id")
    rec.form_type_id = form_type_id
    db.session.commit()
    return jsonify({"success": True})

@verify_bp.route("/export-excel/<int:submission_id>", methods=["GET"])
def export_excel(submission_id):
    records = ExtractedRecord.query.filter_by(
        submission_id=submission_id,
        status='saved'
    ).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Batch {submission_id}"

    if not records:
        ws.append(["No saved records found in this batch"])
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return send_file(output,
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                         as_attachment=True,
                         download_name=f'batch_{submission_id}_export.xlsx')

    # Style constants matching excel_service.py
    HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    HEADER_FONT = Font(color="FFFFFF", bold=True, size=10)
    ALT_ROW_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
    from openpyxl.styles import Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    BORDER = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    # Get enabled fields from form type
    all_fields = []
    for record in records:
        form_type = FormType.query.get(record.form_type_id) if record.form_type_id else None
        if form_type:
            for field in form_type.enabled_fields():
                if not any(f.key == field.key for f in all_fields):
                    all_fields.append(field)

    # Write header using labels
    headers = ['#', 'File', 'Saved At'] + [f.label for f in all_fields]
    ws.append(headers)
    for col_idx, _ in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
    ws.row_dimensions[1].height = 30

    # Write rows
    for row_idx, record in enumerate(records, 1):
        try:
            payload = json.loads(record.raw_extraction or '{}')
            data = payload.get('consensus', payload)
        except Exception:
            data = {}
        row = [
            row_idx,
            record.filename or '',
            record.saved_at.strftime("%Y-%m-%d %H:%M") if record.saved_at else '',
        ]
        for f in all_fields:
            row.append(data.get(f.key, ''))
        ws.append(row)

        # Apply row style
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=row_idx + 1, column=col_idx)
            cell.border = BORDER
            cell.alignment = Alignment(vertical="center")
            if row_idx % 2 == 0:
                cell.fill = ALT_ROW_FILL

    # Auto width
    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=10)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 40)

    ws.freeze_panes = "A2"

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True,
                     download_name=f'batch_{submission_id}_export.xlsx')